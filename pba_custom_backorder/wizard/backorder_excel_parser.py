import base64
import re
import unicodedata
from io import BytesIO

from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError as err:
    openpyxl = None
    _OPENPYXL_ERR = err
else:
    _OPENPYXL_ERR = None

try:
    import xlrd
except ImportError as err:
    xlrd = None
    _XLRD_ERR = err
else:
    _XLRD_ERR = None

FIELD_ALIASES = {
    "item": ("item", "ítem", "linea", "línea", "line", "#"),
    "internal_code": (
        "codigo",
        "código",
        "code",
        "no. parte",
        "no parte",
        "nro parte",
        "numero",
        "número",
        "numero parte",
        "parte",
        "part number",
        "part no",
        "sku",
    ),
    "product_ref": ("referencia", "ref"),
    "description": ("descripcion", "descripción", "description", "nombre", "desc"),
    "supplier_name": ("proveedor", "supplier", "vendor", "adk corporation"),
    "factory": ("fabrica", "fábrica", "factory", "marca"),
    "order_ref": ("no. pedido", "no pedido", "pedido", "order", "po", "proforma"),
    "confirmation": (
        "no. confirmacion venta",
        "no confirmacion venta",
        "confirmacion",
        "confirmación",
        "confirmation",
        "sc",
    ),
    "quantity": ("cantidad a pedir", "cantidad", "qty", "quantity"),
    "uom": ("uni", "uom", "unidad", "um", "st", "kt", "pzs", "pcs"),
    "unit_price": (
        "precio uni",
        "precio uni.",
        "precio unit",
        "precio unitario",
        "precio unit.",
        "costo jpy",
        "costo / costo",
        "costo ult",
        "costo últ",
        "fob japon",
        "fob japan",
        "fob japón",
        "p. uni",
        "p.uni",
        "unit price",
        "unit price.",
        "price uni",
    ),
    "line_total": (
        "total yen",
        "total usd",
        "total eur",
        "total",
        "importe",
        "subtotal",
    ),
}

CURRENCY_HINTS = {
    "yen": "JPY",
    "jpy": "JPY",
    "fob japon": "JPY",
    "fob japan": "JPY",
    "fob japón": "JPY",
    "usd": "USD",
    "dolar": "USD",
    "dólar": "USD",
    "eur": "EUR",
    "euro": "EUR",
}

UOM_SAMPLE_CODES = frozenset(
    {
        "st",
        "kt",
        "pzs",
        "pzs.",
        "pcs",
        "pcs.",
        "set",
        "sets",
        "jgo",
        "jgos",
        "uni",
        "un",
    }
)

HEADER_SCAN_MAX_ROW = 50


def _normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text)


def _match_field(header_norm):
    if not header_norm:
        return None
    for field_key, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if header_norm == alias:
                return field_key
    best_field = None
    best_len = 0
    for field_key, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if len(alias) < 4:
                continue
            if alias in header_norm and len(alias) > best_len:
                best_field = field_key
                best_len = len(alias)
    if best_field:
        return best_field
    for field_key, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if len(alias) >= 4:
                continue
            if header_norm == alias:
                return field_key
    return None


def _refine_quantity_column(headers, mapping):
    norms = [_normalize_header(h) for h in headers]
    for idx, norm in enumerate(norms):
        if "cantidad a pedir" in norm:
            mapping["quantity"] = idx
            return
    for idx, norm in enumerate(norms):
        if norm == "cantidad":
            mapping["quantity"] = idx
            return
    for idx, norm in enumerate(norms):
        if norm in ("cant", "cant."):
            mapping["quantity"] = idx
            return


def _refine_unit_price_column(headers, mapping, currency_code):
    norms = [_normalize_header(h) for h in headers]
    candidates = []
    for idx, header in enumerate(headers):
        norm = norms[idx]
        if "costo" in norm or norm.startswith("fob") or "precio" in norm:
            candidates.append((idx, norm, header))
    if not candidates:
        return
    if currency_code == "USD":
        for idx, norm, header in candidates:
            if "usd" in norm or "ult" in norm or "/" in _cell_str(header):
                mapping["unit_price"] = idx
                return
    if currency_code == "JPY":
        for idx, norm, header in candidates:
            if "jpy" in norm or "yen" in norm or "¥" in _cell_str(header):
                mapping["unit_price"] = idx
                return
    for idx, norm, header in candidates:
        if norm.startswith("fob") or "precio uni" in norm:
            mapping["unit_price"] = idx
            return
    mapping["unit_price"] = candidates[0][0]


def _refine_column_mapping(headers, mapping, sample_rows, currency_code=False):
    _refine_quantity_column(headers, mapping)
    if currency_code:
        _refine_unit_price_column(headers, mapping, currency_code)
    elif "unit_price" not in mapping:
        for col_idx, header in enumerate(headers):
            norm = _normalize_header(header)
            if norm.startswith("fob"):
                mapping["unit_price"] = col_idx
                break
    if "quantity" in mapping and "uom" not in mapping:
        uom_idx = mapping["quantity"] + 1
        if uom_idx < len(headers) and not _cell_str(headers[uom_idx]):
            if _column_looks_like_uom(sample_rows, uom_idx):
                mapping["uom"] = uom_idx
    return mapping


def _column_looks_like_uom(sample_rows, col_idx):
    hits = 0
    checked = 0
    for row in sample_rows:
        if col_idx >= len(row):
            continue
        val = _cell_str(row[col_idx])
        if not val:
            continue
        checked += 1
        if _normalize_header(val) in UOM_SAMPLE_CODES:
            hits += 1
    return checked > 0 and hits >= max(1, checked // 2)


def _detect_currency_from_headers(headers):
    for header in headers:
        norm = _normalize_header(header)
        if not norm:
            continue
        for hint, code in sorted(CURRENCY_HINTS.items(), key=lambda x: -len(x[0])):
            if hint in norm:
                return code
    return False


def _detect_currency_from_filename(filename):
    if not filename:
        return False
    lower = filename.lower()
    for hint, code in sorted(CURRENCY_HINTS.items(), key=lambda x: -len(x[0])):
        if hint in lower:
            return code
    return False


def _confirmation_from_filename(filename):
    if not filename:
        return ""
    match = re.search(r"F-(\d+)-(\d+)", filename, flags=re.IGNORECASE)
    if match:
        return f"F-{match.group(1)}/{match.group(2)}"
    match = re.search(r"(\d{2}-\d{4}-\d{4})", filename, flags=re.IGNORECASE)
    if match:
        return match.group(1)
    match = re.search(r"(\d{4}-\d{2}-\d{2})", filename)
    if match and "reporte" in filename.lower():
        return f"Reporte-{match.group(1)}"
    return ""


def _detect_currency_from_preface_rows(rows_before_header):
    for row in rows_before_header:
        for cell in row:
            text = _cell_str(cell).lower()
            if not text:
                continue
            if "moneda:" in text or "moneda " in text:
                if "usd" in text or "us$" in text or "($)" in text:
                    return "USD"
                if "jpy" in text or "yen" in text or "(¥)" in text:
                    return "JPY"
                if "eur" in text or "euro" in text:
                    return "EUR"
    return False


def _confirmation_from_preface_rows(rows_before_header):
    for row in rows_before_header:
        for cell in row:
            text = _cell_str(cell)
            if not text:
                continue
            match = re.search(r"F-\d+/\d+[A-Z]?", text, flags=re.IGNORECASE)
            if match:
                return match.group(0)
            match = re.search(r"F-\d+-\d+[A-Z]?", text, flags=re.IGNORECASE)
            if match:
                return match.group(0).replace("-", "/", 1)
            match = re.search(r"(\d{2}/\d{2}/\d{4})", text)
            if match and "reporte" in text.lower():
                parts = match.group(1).split("/")
                return f"Reporte-{parts[2]}-{parts[1]}-{parts[0]}"
    return ""


def _normalize_column_mapping(mapping):
    if "part_number" in mapping:
        if "internal_code" not in mapping:
            mapping["internal_code"] = mapping["part_number"]
        del mapping["part_number"]
    return mapping


def _identifier_field_keys(mapping):
    mapping = _normalize_column_mapping(dict(mapping))
    for key in ("internal_code", "product_ref"):
        if key in mapping:
            yield key


def _row_identifier(row_values, mapping):
    for key in _identifier_field_keys(mapping):
        idx = mapping[key]
        if idx < len(row_values):
            val = _cell_str(row_values[idx])
            if val:
                return val
    return ""


def _row_is_data(row_values, mapping):
    if not _row_identifier(row_values, mapping):
        return False
    identifier = _row_identifier(row_values, mapping).lower()
    if identifier in ("total yen", "total usd", "total", "total eur", "numero", "item", "codigo", "código"):
        return False
    item_idx = mapping.get("item")
    if item_idx is not None and item_idx < len(row_values):
        item_val = row_values[item_idx]
        if item_val is not None and str(item_val).strip().lower() in (
            "total yen",
            "total usd",
            "total",
            "item",
        ):
            return False
    desc_idx = mapping.get("description")
    ref_idx = mapping.get("product_ref")
    description = (
        _cell_str(row_values[desc_idx])
        if desc_idx is not None and desc_idx < len(row_values)
        else ""
    )
    product_ref = (
        _cell_str(row_values[ref_idx])
        if ref_idx is not None and ref_idx < len(row_values)
        else ""
    )
    if not description and not product_ref:
        return False
    return True


def _cell_float(value):
    if value is None or value is False:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if not text:
        return 0.0
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _cell_str(value):
    if value is None or value is False:
        return ""
    return str(value).strip()


class _SheetReader:
    def __init__(self, file_kind, workbook):
        self.file_kind = file_kind
        self.workbook = workbook

    @property
    def sheet_names(self):
        if self.file_kind == "xlsx":
            return self.workbook.sheetnames
        return self.workbook.sheet_names()

    def iter_rows(self, sheet_index, min_row, max_row=None):
        if self.file_kind == "xlsx":
            ws = self.workbook.worksheets[sheet_index]
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
                yield list(row)
            return
        sheet = self.workbook.sheet_by_index(sheet_index)
        last = sheet.nrows if max_row is None else min(max_row, sheet.nrows)
        for row_idx in range(min_row - 1, last):
            yield [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]

    def close(self):
        if self.file_kind == "xlsx":
            self.workbook.close()


class BackorderExcelParser:
    @classmethod
    def _decode_file(cls, file_data_b64):
        return base64.b64decode(file_data_b64)

    @classmethod
    def _open_reader(cls, raw, filename=None):
        ext = (filename or "").lower().rsplit(".", 1)[-1] if filename and "." in filename else ""
        if ext == "xls":
            if xlrd is None:
                raise UserError(
                    "La librería xlrd no está instalada. "
                    "Instálela en el entorno de Odoo para importar archivos .xls."
                ) from _XLRD_ERR
            return _SheetReader("xls", xlrd.open_workbook(file_contents=raw))
        if openpyxl is None:
            raise UserError(
                "La librería openpyxl no está instalada. "
                "Instálela en el entorno de Odoo para importar archivos Excel."
            ) from _OPENPYXL_ERR
        return _SheetReader(
            "xlsx",
            openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True),
        )

    @classmethod
    def scan_sheet(cls, reader, sheet_index=0, max_scan_rows=HEADER_SCAN_MAX_ROW):
        best = None
        for row_idx, row in enumerate(
            reader.iter_rows(sheet_index, 1, max_scan_rows),
            start=1,
        ):
            headers = [_cell_str(c) for c in row]
            norms = [_normalize_header(h) for h in headers]
            mapping = {}
            for col_idx, norm in enumerate(norms):
                field_key = _match_field(norm)
                if field_key and field_key not in mapping:
                    mapping[field_key] = col_idx
            mapping = _normalize_column_mapping(mapping)
            if "internal_code" not in mapping:
                continue
            sample_rows = list(reader.iter_rows(sheet_index, row_idx + 1, row_idx + 8))
            preface = list(reader.iter_rows(sheet_index, 1, row_idx - 1))
            currency_code = (
                _detect_currency_from_preface_rows(preface)
                or _detect_currency_from_headers(headers)
            )
            mapping = _normalize_column_mapping(
                _refine_column_mapping(
                    headers, mapping, sample_rows, currency_code=currency_code
                )
            )
            score = len(mapping)
            if "internal_code" in mapping:
                score += 5
            if "quantity" in mapping:
                score += 2
            if "unit_price" in mapping:
                score += 2
            if score > (best["score"] if best else -1):
                best = {
                    "header_row": row_idx,
                    "headers": headers,
                    "column_mapping": mapping,
                    "score": score,
                    "preface_rows": preface,
                    "currency_code": currency_code,
                }
        if not best or "internal_code" not in _normalize_column_mapping(
            best["column_mapping"]
        ):
            sheet_label = reader.sheet_names[sheet_index]
            raise UserError(
                "No se detectó una fila de encabezados válida en la hoja «%s». "
                "Verifique que incluya una columna de número de parte (p. ej. NO. PARTE o NUMERO)."
                % sheet_label
            )
        return best

    @classmethod
    def _pick_best_sheet(cls, reader):
        best_index = 0
        best_scan = None
        for sheet_index in range(len(reader.sheet_names)):
            try:
                scan = cls.scan_sheet(reader, sheet_index=sheet_index)
            except UserError:
                continue
            if best_scan is None or scan["score"] > best_scan["score"]:
                best_scan = scan
                best_index = sheet_index
        if best_scan is None:
            raise UserError(
                "No se encontró ninguna hoja con líneas de producto en el archivo."
            )
        return best_index, best_scan

    @classmethod
    def auto_configure(cls, file_data_b64, filename=None, sheet_index=None):
        raw = cls._decode_file(file_data_b64)
        reader = cls._open_reader(raw, filename=filename)
        try:
            if sheet_index is None:
                sheet_index, scan = cls._pick_best_sheet(reader)
            else:
                scan = cls.scan_sheet(reader, sheet_index=sheet_index)
            headers = scan["headers"]
            currency_code = (
                scan.get("currency_code")
                or _detect_currency_from_preface_rows(scan.get("preface_rows") or [])
                or _detect_currency_from_headers(headers)
                or _detect_currency_from_filename(filename)
            )
            column_mapping = _normalize_column_mapping(
                _refine_column_mapping(
                    headers,
                    dict(scan["column_mapping"]),
                    [],
                    currency_code=currency_code,
                )
            )
            confirmation = _confirmation_from_preface_rows(scan.get("preface_rows") or [])
            if not confirmation:
                confirmation = _confirmation_from_filename(filename)
            conf_idx = scan["column_mapping"].get("confirmation")
            if conf_idx is not None and scan["headers"][conf_idx]:
                confirmation = scan["headers"][conf_idx]
            return {
                "header_row": scan["header_row"],
                "headers": headers,
                "column_mapping": column_mapping,
                "currency_code": currency_code,
                "confirmation_ref": confirmation,
                "sheet_name": reader.sheet_names[sheet_index],
                "sheet_index": sheet_index,
                "sheet_names": reader.sheet_names,
                "file_kind": reader.file_kind,
            }
        finally:
            reader.close()

    @classmethod
    def iter_data_rows(cls, file_data_b64, header_row, column_mapping, filename=None, sheet_index=0):
        raw = cls._decode_file(file_data_b64)
        reader = cls._open_reader(raw, filename=filename)
        try:
            for row in reader.iter_rows(sheet_index, header_row + 1):
                row_values = list(row)
                if not _row_is_data(row_values, column_mapping):
                    continue
                yield cls._row_to_dict(row_values, column_mapping)
        finally:
            reader.close()

    @classmethod
    def _row_to_dict(cls, row_values, column_mapping):
        def get_field(field_key, default=""):
            idx = column_mapping.get(field_key)
            if idx is None or idx >= len(row_values):
                return default
            val = row_values[idx]
            if field_key in ("quantity", "unit_price", "line_total"):
                return _cell_float(val)
            return _cell_str(val)

        internal_code = get_field("internal_code")
        product_ref = get_field("product_ref")
        return {
            "internal_code": internal_code,
            "product_ref": product_ref,
            "description": get_field("description"),
            "supplier_name": get_field("supplier_name"),
            "factory": get_field("factory"),
            "order_ref": get_field("order_ref"),
            "confirmation": get_field("confirmation"),
            "quantity": get_field("quantity", 0.0),
            "uom_name": get_field("uom"),
            "unit_price": get_field("unit_price", 0.0),
            "line_total": get_field("line_total", 0.0),
        }

    @classmethod
    def mapping_to_json(cls, column_mapping, headers):
        result = {}
        for field_key, col_idx in column_mapping.items():
            if col_idx < len(headers):
                label = headers[col_idx]
                if not label and field_key == "uom":
                    label = "(columna sin título / UdM)"
                result[field_key] = label
        return result

    @classmethod
    def json_to_column_mapping(cls, mapping_json, headers):
        norm_headers = {_normalize_header(h): i for i, h in enumerate(headers)}
        column_mapping = {}
        for field_key, header_name in (mapping_json or {}).items():
            if not header_name:
                continue
            if header_name.startswith("(columna"):
                if field_key == "uom":
                    qty_idx = column_mapping.get("quantity")
                    if qty_idx is None:
                        for fk, col in norm_headers.items():
                            if _match_field(fk) == "quantity":
                                qty_idx = col
                                break
                    if qty_idx is not None:
                        column_mapping["uom"] = qty_idx + 1
                continue
            norm = _normalize_header(header_name)
            if norm in norm_headers:
                column_mapping[field_key] = norm_headers[norm]
            else:
                for idx, header in enumerate(headers):
                    if _normalize_header(header) == norm:
                        column_mapping[field_key] = idx
                        break
        return _normalize_column_mapping(column_mapping)
